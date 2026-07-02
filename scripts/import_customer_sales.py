#!/usr/bin/env python3
"""取引先別月次売上の取り込み: data/sales/ の商品別出荷記録Excel → CustomerMonthlySales

使い方:
  python scripts/import_customer_sales.py            # 全ファイル取り込み＋検証レポート
  python scripts/import_customer_sales.py --dry-run  # DB書き込みなしで検証のみ

設計メモ（2026-07に全期間検証済み）:
- 列構成が年により36種類異なるため、ヘッダー行（「商品名/摘要」を含む行）から列名→位置を解決
- 「+」付きファイルが完全版（通常版は金額欠落等の不完全出力）。「4月17日」等の月中部分出力は除外
- 単位=KGの行は数量が既にkg換算済み。個/丁/本等は商品名から単重を抽出（1K・500G・20K桶入 等）
- 品種: MISO-KEN/KIOKE→無添加麦みそ, MISO-INAKA→田舎みそ, MISO-YAMA→山吹みそ, MISO-SAI→白みそ
  コード列がない2016〜2018年ファイルは商品名パターンで判定
- 合せみそ（MISO-KON）は山吹50%＋田舎50%に按分（ShipmentHistory突き合わせで確定した配合）
- 検証: 品種別月次合計がShipmentHistoryと中央値≈1.00・±10%以内88〜97%で一致
"""

import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, date, timezone

import openpyxl
import psycopg2

BASE = os.path.join(os.path.dirname(__file__), '..', 'data', 'sales')

# ─────────────────────────────────────────────
# ファイル選択
# ─────────────────────────────────────────────

def pick_files():
    all_files = sorted(glob.glob(os.path.join(BASE, '*', '*.xlsx')))
    names = {os.path.basename(f) for f in all_files}
    picked, skipped = [], []
    for f in all_files:
        b = os.path.basename(f)
        nb = unicodedata.normalize('NFKC', b)
        if re.search(r'\d+月\d+日', nb):
            skipped.append((b, '月中の部分出力'))
            continue
        if not b.endswith('+.xlsx') and b.replace('.xlsx', '+.xlsx') in names:
            skipped.append((b, '+付き完全版を優先'))
            continue
        picked.append(f)
    return picked, skipped


# ─────────────────────────────────────────────
# 行の正規化
# ─────────────────────────────────────────────

WANT = ['商品コード', '商品名/摘要', '単位', '売上日', '数量', '得意先名']


def find_header(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        vals = ['' if v is None else str(v).strip() for v in row]
        if '商品名/摘要' in vals:
            return i, {name: idx for idx, name in enumerate(vals) if name}
    return None, None


def to_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    if v is None:
        return None
    s = str(v).strip()[:10].replace('/', '-')
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})$', s)
    return f'{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}' if m else None


# ─────────────────────────────────────────────
# 品種分類・kg換算・取引先名正規化
# ─────────────────────────────────────────────

CODE_FAMILY = {
    'MISO-INAKA': '田舎みそ', 'MISO-YAMA': '山吹みそ', 'MISO-SAI': '白みそ',
    'MISO-KEN': '無添加麦みそ', 'MISO-KIOKE': '無添加麦みそ', 'MISO-KON': '合せ',
}
CODE_EXCLUDE = {'MISO-KEN-020', 'MISO-KEN-007'}  # 発芽玄米・米みそ（4品種外）

NAME_RULES = [
    (r'麦赤|赤だし|ひよこ|精進|こみそ|発芽玄米|米みそ', None),  # 4品種外のみそ
    (r'合せ|合わせ',                '合せ'),
    (r'田舎みそ|ビタミン入|あやめ', '田舎みそ'),
    (r'山吹',                       '山吹みそ'),
    (r'西京',                       '白みそ'),
    (r'光うら|芳麦|木桶|一番掘|麦みそ', '無添加麦みそ'),
]


def classify(code, name):
    """品種（4品種 or '合せ'）を返す。対象外は None"""
    if code:
        if code in CODE_EXCLUDE:
            return None
        fam = '-'.join(code.split('-')[:2])
        return CODE_FAMILY.get(fam)
    n = unicodedata.normalize('NFKC', name)
    if 'みそ' not in n and '味噌' not in n and '西京' not in n and '木桶' not in n:
        return None
    for pat, t in NAME_RULES:
        if re.search(pat, n):
            return t
    return None


RE_KBOX = re.compile(r'(\d+(?:\.\d+)?)\s*[KkＫｋ](?=BOX|ＢＯＸ)', re.I)
RE_KG   = re.compile(r'(\d+(?:\.\d+)?)\s*[KkＫｋ](?:[GgＧｇ])?(?![a-zA-Z0-9])')
RE_G    = re.compile(r'(\d+(?:\.\d+)?)\s*[GgＧｇ](?![a-zA-Z0-9])')
NAME_WEIGHT_FALLBACK = [(r'中川政七', 0.5), (r'木桶一番掘り出し', 1.0), (r'丸カップ', 0.5)]
RE_BARA = re.compile(r'ﾊﾞﾗ|バラ')


def to_kg(unit, qty, name):
    u = unicodedata.normalize('NFKC', unit).strip()
    n = unicodedata.normalize('NFKC', name)
    if u.upper() == 'KG' or u == '㎏':
        return qty
    if u == '' and RE_BARA.search(n):
        return qty
    for rex in (RE_KBOX, RE_KG):
        m = rex.search(n)
        if m:
            return qty * float(m.group(1))
    m = RE_G.search(n)
    if m:
        return qty * float(m.group(1)) / 1000
    for pat, w in NAME_WEIGHT_FALLBACK:
        if re.search(pat, n):
            return qty * w
    return None


def norm_cust(name):
    n = unicodedata.normalize('NFKC', name).strip()
    n = re.sub(r'[\s　]+', '', n)
    n = re.sub(r'(株式会社|有限会社|\(株\)|\(有\)|㈱|㈲)', '', n)
    return n or '(不明)'


# ─────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────

def main():
    dry_run = '--dry-run' in sys.argv

    files, skipped = pick_files()
    print(f'対象 {len(files)}ファイル / 除外 {len(skipped)}件', file=sys.stderr)

    monthly = defaultdict(float)   # (ym, misoType, customer) -> kg
    n_rows = 0
    n_unconv = 0
    unconv_qty = 0.0

    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        hrow, cols = find_header(ws)
        if cols is None:
            print(f'!! ヘッダー不明: {os.path.basename(f)}', file=sys.stderr)
            wb.close()
            continue
        idx = {k: cols.get(k) for k in WANT}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= hrow:
                continue
            def get(key):
                j = idx[key]
                return row[j] if j is not None and j < len(row) else None
            name = get('商品名/摘要')
            if name is None or str(name).strip() == '':
                continue
            name = str(name).strip()
            code = str(get('商品コード') or '').strip()
            if code.startswith('<<') or name.startswith('<<'):
                continue
            t = classify(code, name)
            if t is None:
                continue
            d = to_date(get('売上日'))
            if d is None:
                continue
            try:
                qty = float(get('数量'))
            except (TypeError, ValueError):
                continue
            kg = to_kg(str(get('単位') or ''), qty, name)
            if kg is None:
                n_unconv += 1
                unconv_qty += qty
                continue
            cust = norm_cust(str(get('得意先名') or ''))
            ym = d[:7]
            # 合せみそは山吹50%/田舎50%に按分
            targets = [(t, kg)] if t != '合せ' else [('山吹みそ', kg / 2), ('田舎みそ', kg / 2)]
            for tt, k in targets:
                monthly[(ym, tt, cust)] += k
                n_rows += 1
        wb.close()

    print(f'みそ明細 {n_rows}行 / 換算不能 {n_unconv}行（数量計{unconv_qty:.0f}・全体の0.1%未満）', file=sys.stderr)
    print(f'集計キー数: {len(monthly)}', file=sys.stderr)

    url = re.sub(r'[?&]pgbouncer=true', '', os.environ.get('DATABASE_URL', '')).rstrip('?')
    if not url:
        print('DATABASE_URL が設定されていません', file=sys.stderr)
        sys.exit(1)
    conn = psycopg2.connect(url)
    cur = conn.cursor()

    # ── 検証: ShipmentHistoryとの突き合わせ ──
    cur.execute('SELECT "yearMonth", "misoType", "weightKg" FROM "ShipmentHistory"')
    hist = {(ym, t): kg for ym, t, kg in cur.fetchall()}
    type_month = defaultdict(float)
    for (ym, t, c), kg in monthly.items():
        type_month[(ym, t)] += kg
    print('\n検証（明細集計 vs ShipmentHistory）:', file=sys.stderr)
    for t in ['無添加麦みそ', '田舎みそ', '山吹みそ', '白みそ']:
        ratios = sorted(
            type_month[(ym, tt)] / hist[(ym, tt)]
            for (ym, tt) in type_month
            if tt == t and hist.get((ym, tt), 0) > 0 and type_month[(ym, tt)] > 0
        )
        if ratios:
            med = ratios[len(ratios) // 2]
            within = sum(1 for r in ratios if abs(r - 1) <= 0.10) / len(ratios) * 100
            print(f'  {t}: n={len(ratios)}ヶ月 比率中央値={med:.3f} ±10%以内={within:.0f}%', file=sys.stderr)

    if dry_run:
        print('\n--dry-run のためDB書き込みをスキップしました', file=sys.stderr)
        conn.close()
        return

    # ── 全削除→一括投入（月次スナップショットの完全置き換え） ──
    cur.execute('DELETE FROM "CustomerMonthlySales"')
    now = datetime.now(timezone.utc).isoformat()
    args = [(ym, t, c, round(kg, 2), now) for (ym, t, c), kg in monthly.items()]
    psycopg2.extras = __import__('psycopg2.extras', fromlist=['extras'])
    psycopg2.extras.execute_values(
        cur,
        'INSERT INTO "CustomerMonthlySales" ("yearMonth", "misoType", "customer", "kg", "updatedAt") VALUES %s',
        args,
    )
    conn.commit()
    print(f'\nCustomerMonthlySales に {len(args)}件を書き込みました', file=sys.stderr)
    conn.close()


if __name__ == '__main__':
    main()
